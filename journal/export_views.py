from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from core.models import Class, Subject
from accounts.models import User
from .models import Grade, Attendance
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

@login_required
def export_grades_excel(request):
    """Export grades to Excel file"""
    if request.user.role not in ['teacher', 'director']:
        messages.error(request, "Ruxsat yo'q!")
        return redirect('home')
    
    class_id = request.GET.get('class_id')
    subject_id = request.GET.get('subject_id')
    
    if not class_id or not subject_id:
        messages.error(request, "Sinf va fanni tanlang!")
        return redirect('gradebook')
    
    selected_class = get_object_or_404(Class, id=class_id)
    selected_subject = get_object_or_404(Subject, id=subject_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Baholar"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = f"Baholar Hisoboti - {selected_class.name} - {selected_subject.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Sana: {datetime.now().strftime('%d.%m.%Y')}"
    ws.merge_cells('A2:D2')
    
    # Headers
    headers = ['#', "O'quvchi", 'Baho', 'Sana']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    grades = Grade.objects.filter(
        student__student_class=selected_class,
        subject=selected_subject
    ).select_related('student').order_by('student__last_name', 'date')
    
    row = 5
    for idx, grade in enumerate(grades, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f"{grade.student.first_name} {grade.student.last_name}")
        ws.cell(row=row, column=3, value=grade.value)
        ws.cell(row=row, column=4, value=grade.date.strftime('%d.%m.%Y'))
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"baholar_{selected_class.name}_{selected_subject.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_attendance_excel(request):
    """Export attendance to Excel file"""
    if request.user.role not in ['teacher', 'director']:
        messages.error(request, "Ruxsat yo'q!")
        return redirect('home')
    
    class_id = request.GET.get('class_id')
    
    if not class_id:
        messages.error(request, "Sinfni tanlang!")
        return redirect('attendance')
    
    selected_class = get_object_or_404(Class, id=class_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"
    
    # Header styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = f"Davomat Hisoboti - {selected_class.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Sana: {datetime.now().strftime('%d.%m.%Y')}"
    ws.merge_cells('A2:D2')
    
    # Headers
    headers = ['#', "O'quvchi", 'Status', 'Sana']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    attendance_records = Attendance.objects.filter(
        student__student_class=selected_class
    ).select_related('student').order_by('date', 'student__last_name')
    
    status_map = {
        'present': 'Keldi',
        'absent': 'Kelmadi',
        'late': 'Kech qoldi',
        'excused': 'Sababli'
    }
    
    row = 5
    for idx, record in enumerate(attendance_records, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=f"{record.student.first_name} {record.student.last_name}")
        ws.cell(row=row, column=3, value=status_map.get(record.status, record.status))
        ws.cell(row=row, column=4, value=record.date.strftime('%d.%m.%Y'))
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"davomat_{selected_class.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_students_excel(request):
    """Export students list to Excel file"""
    if request.user.role not in ['teacher', 'director']:
        messages.error(request, "Ruxsat yo'q!")
        return redirect('home')
    
    class_id = request.GET.get('class_id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"
    
    # Header styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    # Title
    if class_id:
        selected_class = get_object_or_404(Class, id=class_id)
        ws['A1'] = f"O'quvchilar Ro'yxati - {selected_class.name}"
        students = User.objects.filter(role='student', student_class=selected_class)
    else:
        ws['A1'] = "O'quvchilar Ro'yxati - Barcha Sinflar"
        students = User.objects.filter(role='student')
    
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Sana: {datetime.now().strftime('%d.%m.%Y')}"
    ws.merge_cells('A2:F2')
    
    # Headers
    headers = ['#', 'Ism', 'Familiya', 'Username', 'Email', 'Sinf']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    for idx, student in enumerate(students.order_by('student_class', 'last_name'), 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=student.first_name)
        ws.cell(row=row, column=3, value=student.last_name)
        ws.cell(row=row, column=4, value=student.username)
        ws.cell(row=row, column=5, value=student.email or '')
        ws.cell(row=row, column=6, value=str(student.student_class) if student.student_class else '')
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    class_suffix = f"_{selected_class.name}" if class_id else "_barcha"
    filename = f"oquvchilar{class_suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
