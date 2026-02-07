from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from accounts.models import User
from core.models import Class
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

@login_required
def download_students_template(request):
    """Download Excel template for student import"""
    if request.user.role != 'director':
        messages.error(request, "Ruxsat yo'q!")
        return redirect('home')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = "O'quvchilar Import Template"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Quyidagi formatda to'ldiring. Birinchi qatorni o'zgartirmang!"
    ws.merge_cells('A2:F2')
    
    # Headers
    headers = ['Ism*', 'Familiya*', 'Username*', 'Email', 'Parol*', 'Sinf ID*']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Example data
    ws.cell(row=5, column=1, value="Ali")
    ws.cell(row=5, column=2, value="Valiyev")
    ws.cell(row=5, column=3, value="ali.valiyev")
    ws.cell(row=5, column=4, value="ali@example.com")
    ws.cell(row=5, column=5, value="parol123")
    ws.cell(row=5, column=6, value="1")
    
    # Instructions
    ws['A7'] = "Izoh: * belgisi majburiy maydonlarni bildiradi"
    ws['A8'] = "Sinf ID: Tizimda mavjud sinf ID raqamini kiriting"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"oquvchilar_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def import_students(request):
    """Import students from Excel file"""
    if request.user.role != 'director':
        messages.error(request, "Ruxsat yo'q!")
        return redirect('home')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Start from row 5 (after headers and example)
            for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
                # Skip empty rows
                if not any(row):
                    continue
                
                try:
                    first_name = row[0]
                    last_name = row[1]
                    username = row[2]
                    email = row[3] or ''
                    password = row[4]
                    class_id = row[5]
                    
                    # Validate required fields
                    if not all([first_name, last_name, username, password, class_id]):
                        errors.append(f"Qator {row_num}: Majburiy maydonlar to'ldirilmagan")
                        error_count += 1
                        continue
                    
                    # Check if username already exists
                    if User.objects.filter(username=username).exists():
                        errors.append(f"Qator {row_num}: '{username}' allaqachon mavjud")
                        error_count += 1
                        continue
                    
                    # Get class
                    try:
                        student_class = Class.objects.get(id=int(class_id))
                    except (Class.DoesNotExist, ValueError):
                        errors.append(f"Qator {row_num}: Sinf ID {class_id} topilmadi")
                        error_count += 1
                        continue
                    
                    # Create student
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        role='student',
                        student_class=student_class
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Qator {row_num}: Xatolik - {str(e)}")
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f"{success_count} ta o'quvchi muvaffaqiyatli yuklandi!")
            if error_count > 0:
                messages.warning(request, f"{error_count} ta xatolik. Quyida batafsil:")
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
            
            return redirect('import_students')
            
        except Exception as e:
            messages.error(request, f"Fayl o'qishda xatolik: {str(e)}")
            return redirect('import_students')
    
    # GET request - show form
    classes = Class.objects.all()
    return render(request, 'core/import_students.html', {'classes': classes})
